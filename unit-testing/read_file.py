from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl



path = "./Unit Testing.xlsx"

unit_testing = openpyxl.load_workbook(path)
login_testing = unit_testing["Login"]


columns = login_testing.max_column
rows = login_testing.max_row

exception_case = [rows-5, rows-6]

for c in range(6, 7):
    for r in range(14, rows-4):
        value = login_testing.cell(row=r, column=c).value

        if (value== "O"):
            print(login_testing.cell(row=r, column=4).value)


